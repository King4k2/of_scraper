import pickle
from selenium.webdriver.chrome.webdriver import WebDriver as Chrome
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
import time
import openpyxl
import json



def onlyfans_login():
    wb = openpyxl.load_workbook(filename="OnlyFansConfig.xlsx")
    sh = wb["Sheet1"]  # A - account name ; B - user-agent ; C - sec-ch-ua ; D - cookie
    acc_name = input("input OnlyFans Account name: ")
    a = 0
    for i in range(1, sh.max_row+1):
        if sh[f"A{i}"].value == acc_name:
            a = i
            break
    if a == 0:
        if sh.max_row == 1:
            a = sh.max_row
        else:
            a = sh.max_row+1
    options = ChromeOptions()
    service = ChromeService(executable_path='chromedriver-win64/chromedriver.exe')
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36")
    options.add_argument(r'user-data-dir=C:\Users\King\AppData\Local\Google\Chrome\User Data')
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument('--profile-directory=Profile 1')
    driver = Chrome(service=service, options=options)
    driver.get("https://onlyfans.com/")
    # Save cookie
    #i = input("Input something: ")
    #cookies_ = driver.get_cookies()
    #with open(f'{acc_name}_OFS_cookies.pkl', 'wb')as file:
        #pickle.dump(cookies_, file)

    # Load cookie
    for cookie in pickle.load(open(f'{acc_name}_OFS_cookies.pkl', 'rb')):
        driver.add_cookie(cookie)
        time.sleep(5)
        driver.refresh()
        time.sleep(100)
    driver.quit()
    # res = req_interceptor(driver)


def req_interceptor(driver):
    for request in driver.requests:
        if request.url != "https://onlyfans.com/api2/v2/users/me":
            continue
        print(request)
        if request.headers is not None:
            print(request.headers)
            print(request.headers.get("sec-ch-ua"))
            print(request.headers.get("user-agent"))
            print(request.headers.get("cookie"))

            print("\n-----------------\n")
            return [request.headers.get("user-agent"), request.headers.get("accept-language"),
                    request.headers.get("sec-ch-ua"), request.headers.get("cookie")]
    return "Something went wrong. Try again"


if __name__ == "__main__":
    onlyfans_login()
